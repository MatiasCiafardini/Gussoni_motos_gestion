from __future__ import annotations

import re
import unicodedata
from io import BytesIO
from typing import Any, Callable, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.domain_constants import EstadoStock, TipoMovimientoStock
from app.core.catalog_cache import CatalogCache
from app.data.database import SessionLocal
from app.repositories.vehiculos_repository import VehiculosRepository
from app.services.audit_log_service import AuditLogService
from app.services.catalogos_service import CatalogosService
from app.services.stock_service import StockService


SYSTEM_FIELDS = [
    "marca",
    "modelo",
    "anio",
    "nro_certificado",
    "nro_dnrpa",
    "lca",
    "numero_cuadro",
    "numero_motor",
    "precio_lista",
    "color_id",
    "observaciones",
]


FIELD_LABELS = {
    "marca": "Marca",
    "modelo": "Modelo",
    "anio": "Año modelo",
    "nro_certificado": "N° certificado",
    "nro_dnrpa": "N° DNRPA",
    "lca": "LCA",
    "numero_cuadro": "N° cuadro",
    "numero_motor": "N° motor",
    "precio_lista": "Precio lista",
    "color_id": "Color",
    "observaciones": "Observaciones",
}


HEADER_ALIASES = {
    "modelo": {"modelo", "modelo unidad"},
    "nro_certificado": {"n certificado", "nro certificado", "numero certificado", "certificado"},
    "nro_dnrpa": {"n dnrpa", "nro dnrpa", "numero dnrpa", "dnrpa"},
    "numero_cuadro": {"n cuadro", "nro cuadro", "numero cuadro", "cuadro", "chasis", "n chasis"},
    "numero_motor": {"n motor", "nro motor", "numero motor", "motor"},
    "lca": {"lca", "expediente", "if", "expediente if", "licencia configuracion ambiental"},
    "anio": {"ano modelo", "año modelo", "anio modelo", "ano", "año", "anio"},
    "color": {"color"},
    "ubicacion": {"ubicacion", "ubicacion stock", "ubicacio n"},
    "remito": {"remito"},
    "factura": {"factura"},
    "precio_lista": {"precio total", "precio lista", "precio", "importe"},
}


KNOWN_BRANDS = {
    "MOTOMEL",
    "ZANELLA",
    "MONDIAL",
    "CORVEN",
    "SUZUKI",
    "DRAGON",
    "BRAVA",
    "GAF",
    "HONDA",
    "YAMAHA",
    "BAJAJ",
}

NON_BRAND_SHEETS = {
    "FACTURA",
    "FACTURAS",
    "STOCK",
    "RESUMEN",
    "DATOS",
    "CLIENTES",
}


class ImportacionCertificadosService:
    """Importacion avanzada de certificados desde excels reales con preview."""

    def __init__(self) -> None:
        self._catalogos = CatalogosService()
        self._stock = StockService()
        self._audit = AuditLogService()

    def generar_preview(self, file_bytes: bytes) -> Dict[str, Any]:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        color_map = self._color_map()

        rows: List[Dict[str, Any]] = []
        errores: List[str] = []

        with SessionLocal() as db:
            known_brands = self._known_brands(db)
            for ws in wb.worksheets:
                try:
                    if self._is_informative_sheet(ws.title):
                        continue

                    header_row, mapping = self._detect_header(ws)
                    if not header_row:
                        errores.append(f"Hoja {ws.title}: no se detectaron encabezados de vehiculos.")
                        continue

                    marca = self._detect_brand(ws, known_brands)
                    for excel_row in range(header_row + 1, ws.max_row + 1):
                        parsed = self._parse_row(ws, excel_row, mapping, marca, color_map)
                        if not parsed:
                            continue
                        rows.append(self._compare_row(db, parsed))
                except Exception as exc:
                    errores.append(f"Hoja {ws.title}: {exc}")

        summary = {
            "total": len(rows),
            "nuevos": sum(1 for r in rows if r["estado"] == "NUEVO"),
            "existentes_iguales": sum(1 for r in rows if r["estado"] == "EXISTENTE_IGUAL"),
            "existentes_diferencias": sum(1 for r in rows if r["estado"] == "EXISTENTE_DIFERENCIAS"),
            "ambiguos": sum(1 for r in rows if r["estado"] == "DUPLICADO_AMBIGUO"),
            "errores": sum(1 for r in rows if r["estado"] == "ERROR"),
        }
        return {"success": True, "rows": rows, "errores": errores, "summary": summary}

    def aplicar(
        self,
        rows: List[Dict[str, Any]],
        progress_callback: Optional[Callable[[int, int, Dict[str, Any]], None]] = None,
    ) -> Dict[str, Any]:
        creados = 0
        actualizados = 0
        omitidos = 0
        errores: List[str] = []
        total = len(rows)

        db: Session = SessionLocal()
        repo = VehiculosRepository(db)
        try:
            self._crear_colores_faltantes_en_sesion(db, rows)
            for index, row in enumerate(rows, start=1):
                if progress_callback:
                    progress_callback(index, total, row)
                action = row.get("accion")
                payload = row.get("payload") or {}
                existing = row.get("existing") or {}
                source = {
                    "hoja": row.get("hoja"),
                    "fila": row.get("fila"),
                    "importacion": "certificados_avanzada",
                }

                try:
                    if action in (None, "", "OMITIR"):
                        omitidos += 1
                        continue

                    if action == "CREAR":
                        self._validate_create_payload(payload, row)
                        new_id = repo.create_vehiculo(payload)
                        self._stock.registrar_movimiento(
                            db,
                            vehiculo_id=int(new_id),
                            tipo_movimiento=TipoMovimientoStock.INGRESO,
                            estado_anterior_id=None,
                            estado_nuevo_id=payload.get("estado_stock_id"),
                            origen_tipo="importacion_certificados",
                            origen_id=None,
                            observaciones=f"Alta desde hoja {row.get('hoja')} fila {row.get('fila')}.",
                        )
                        self._audit.registrar(
                            db,
                            entidad="vehiculos",
                            entidad_id=int(new_id),
                            accion="IMPORTACION_CERTIFICADOS_CREAR",
                            datos_nuevos=payload,
                            contexto=source,
                        )
                        creados += 1
                        continue

                    if action in ("COMPLETAR_VACIOS", "ACTUALIZAR_DIFERENCIAS"):
                        vehiculo_id = existing.get("id")
                        if not vehiculo_id:
                            raise ValueError("No hay vehiculo existente asociado.")
                        update_data = self._build_update_payload(row, only_empty=(action == "COMPLETAR_VACIOS"))
                        if not update_data:
                            omitidos += 1
                            continue
                        repo.update(int(vehiculo_id), update_data)
                        self._audit.registrar(
                            db,
                            entidad="vehiculos",
                            entidad_id=int(vehiculo_id),
                            accion=self._audit_action_for_import(action),
                            datos_previos={k: existing.get(k) for k in update_data},
                            datos_nuevos=update_data,
                            contexto=source,
                        )
                        actualizados += 1
                        continue

                    omitidos += 1
                except Exception as exc:
                    errores.append(f"Hoja {row.get('hoja')} fila {row.get('fila')}: {exc}")

            if errores:
                db.rollback()
                return {"success": False, "errores": errores}

            db.commit()
            CatalogCache.get().invalidate("colores")
            return {
                "success": True,
                "creados": creados,
                "actualizados": actualizados,
                "omitidos": omitidos,
            }
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()

    def colores_faltantes(self, rows: List[Dict[str, Any]]) -> List[str]:
        missing = set()
        for row in rows:
            excel = row.get("excel") or {}
            payload = row.get("payload") or {}
            color = _clean_str(excel.get("color_nombre"))
            if color and _is_empty(payload.get("color_id")):
                missing.add(color.strip().upper())
        return sorted(missing)

    def crear_colores_faltantes(self, colores: List[str]) -> Dict[str, Any]:
        creados = []
        existentes = []
        normalizados = []
        for color in colores:
            value = _clean_str(color).upper()
            if value and value not in normalizados:
                normalizados.append(value)

        if not normalizados:
            return {"success": True, "creados": [], "existentes": []}

        db: Session = SessionLocal()
        try:
            for color in normalizados:
                existente = db.execute(
                    text(
                        """
                        SELECT id, nombre
                        FROM colores
                        WHERE UPPER(nombre) = :nombre
                        LIMIT 1
                        """
                    ),
                    {"nombre": color.upper()},
                ).mappings().first()
                if existente:
                    existentes.append(dict(existente))
                    continue

                res = db.execute(
                    text("INSERT INTO colores (nombre) VALUES (:nombre)"),
                    {"nombre": color},
                )
                color_id = int(res.lastrowid or 0)
                creados.append({"id": color_id, "nombre": color})
                self._audit.registrar(
                    db,
                    entidad="colores",
                    entidad_id=color_id or None,
                    accion="IMPORTACION_CERTIFICADOS_CREAR_COLOR",
                    datos_nuevos={"nombre": color},
                    contexto={"importacion": "certificados_avanzada"},
                )

            db.commit()
            CatalogCache.get().invalidate("colores")
            return {"success": True, "creados": creados, "existentes": existentes}
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()

    def _detect_header(self, ws) -> Tuple[Optional[int], Dict[str, int]]:
        best_row = None
        best_mapping: Dict[str, int] = {}
        best_score = 0

        max_scan = min(ws.max_row, 30)
        for row_idx in range(1, max_scan + 1):
            mapping: Dict[str, int] = {}
            for col_idx in range(1, ws.max_column + 1):
                header = _norm(ws.cell(row=row_idx, column=col_idx).value)
                if not header:
                    continue
                field = self._map_header(header)
                if field and field not in mapping:
                    mapping[field] = col_idx
            score = len(mapping)
            if score > best_score:
                best_score = score
                best_row = row_idx
                best_mapping = mapping

        if best_score < 5:
            return None, {}
        return best_row, best_mapping

    def _map_header(self, header: str) -> Optional[str]:
        for field, aliases in HEADER_ALIASES.items():
            if header in aliases:
                return field
        for field, aliases in HEADER_ALIASES.items():
            if any(alias and alias in header for alias in aliases):
                return field
        return None

    def _known_brands(self, db: Session) -> set[str]:
        brands = set(KNOWN_BRANDS)
        rows = db.execute(
            text(
                """
                SELECT DISTINCT marca
                FROM vehiculos
                WHERE marca IS NOT NULL
                  AND TRIM(marca) <> ''
                """
            )
        ).mappings().all()
        for row in rows:
            brand = _clean_str(row.get("marca")).upper()
            if brand:
                brands.add(brand)
        return brands

    def _is_informative_sheet(self, sheet_name: str) -> bool:
        normalized = _norm(sheet_name).upper()
        return normalized in NON_BRAND_SHEETS

    def _detect_brand(self, ws, known_brands: set[str]) -> str:
        sheet_brand = _brand_from_text(ws.title, known_brands)
        if sheet_brand:
            return sheet_brand

        sheet_name = _clean_str(ws.title).upper()
        if sheet_name and _norm(sheet_name).upper() not in NON_BRAND_SHEETS:
            return sheet_name

        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            text = " ".join(str(v or "") for v in row)
            brand = _brand_from_text(text, known_brands)
            if brand:
                return brand
        return ""

    def _parse_row(
        self,
        ws,
        excel_row: int,
        mapping: Dict[str, int],
        marca: str,
        color_map: Dict[str, Dict[str, Any]],
    ) -> Optional[Dict[str, Any]]:
        raw: Dict[str, Any] = {}
        for field, col_idx in mapping.items():
            raw[field] = ws.cell(row=excel_row, column=col_idx).value

        completed_columns = sum(1 for v in raw.values() if _clean_str(v))
        if completed_columns <= 2:
            return None

        core_values = [
            raw.get("modelo"),
            raw.get("nro_certificado"),
            raw.get("nro_dnrpa"),
            raw.get("numero_cuadro"),
            raw.get("numero_motor"),
        ]
        if not any(_clean_str(v) for v in core_values):
            return None

        color_name = _clean_str(raw.get("color"))
        color_obj = color_map.get(_norm_color(color_name)) if color_name else None
        observaciones = self._build_observaciones(raw)

        payload = {
            "marca": marca or None,
            "modelo": _clean_str(raw.get("modelo")) or None,
            "anio": _parse_int(raw.get("anio")),
            "nro_certificado": _clean_str(raw.get("nro_certificado")) or None,
            "nro_dnrpa": _clean_str(raw.get("nro_dnrpa")) or None,
            "lca": _clean_str(raw.get("lca")) or None,
            "numero_cuadro": _clean_str(raw.get("numero_cuadro")) or None,
            "numero_motor": _clean_str(raw.get("numero_motor")) or None,
            "precio_lista": _parse_money(raw.get("precio_lista")),
            "color_id": color_obj.get("id") if color_obj else None,
            "estado_stock_id": EstadoStock.DISPONIBLE,
            "estado_moto_id": self._default_condicion_id(),
            "observaciones": observaciones or None,
        }

        return {
            "hoja": ws.title,
            "fila": excel_row,
            "payload": payload,
            "excel": {
                **payload,
                "color_nombre": color_name,
                "ubicacion": _clean_str(raw.get("ubicacion")),
                "remito": _clean_str(raw.get("remito")),
                "factura": _clean_str(raw.get("factura")),
            },
        }

    def _compare_row(self, db: Session, parsed: Dict[str, Any]) -> Dict[str, Any]:
        payload = parsed["payload"]
        errors = self._row_errors(payload, parsed.get("excel") or {})
        matches = self._find_existing(db, payload)

        base = {
            "hoja": parsed["hoja"],
            "fila": parsed["fila"],
            "payload": payload,
            "excel": parsed["excel"],
            "existing": None,
            "diferencias": [],
            "estado": "NUEVO",
            "accion_sugerida": "CREAR",
            "accion": "CREAR",
            "mensaje": "",
        }

        if len(matches) > 1:
            base.update({
                "estado": "DUPLICADO_AMBIGUO",
                "accion_sugerida": "OMITIR",
                "accion": "OMITIR",
                "mensaje": f"Coinciden {len(matches)} vehiculos. Revisar manualmente.",
            })
            return base

        if errors:
            base.update({"estado": "ERROR", "accion_sugerida": "OMITIR", "accion": "OMITIR", "mensaje": "; ".join(errors)})
            return base

        if not matches:
            create_errors = self._create_payload_errors(payload, parsed)
            if create_errors:
                base.update({
                    "estado": "ERROR",
                    "accion_sugerida": "OMITIR",
                    "accion": "OMITIR",
                    "mensaje": "; ".join(create_errors),
                })
                return base

            color_name = (parsed.get("excel") or {}).get("color_nombre")
            if color_name and _is_empty(payload.get("color_id")):
                base.update({
                    "estado": "NUEVO",
                    "accion_sugerida": "CREAR",
                    "accion": "CREAR",
                    "mensaje": f"Color pendiente de agregar al catalogo: {color_name}.",
                    "diferencias": [
                        {
                            "campo": "color_id",
                            "label": FIELD_LABELS["color_id"],
                            "sistema": None,
                            "excel": color_name,
                        }
                    ],
                })
            return base

        existing = matches[0]
        diffs = self._diffs(existing, payload)
        color_name = (parsed.get("excel") or {}).get("color_nombre")
        if color_name and _is_empty(payload.get("color_id")):
            diffs.append(
                {
                    "campo": "color_id",
                    "label": FIELD_LABELS["color_id"],
                    "sistema": existing.get("color_nombre") or existing.get("color_id"),
                    "excel": color_name,
                }
            )
        if not diffs:
            estado = "EXISTENTE_IGUAL"
            accion = "OMITIR"
            msg = "Sin diferencias."
        else:
            estado = "EXISTENTE_DIFERENCIAS"
            accion = "ACTUALIZAR_DIFERENCIAS"
            msg = f"{len(diffs)} diferencia(s)."

        base.update({
            "existing": existing,
            "diferencias": diffs,
            "estado": estado,
            "accion_sugerida": accion,
            "accion": accion,
            "mensaje": msg,
        })
        return base

    def _find_existing(self, db: Session, payload: Dict[str, Any]) -> List[Dict[str, Any]]:
        checks = {
            "numero_motor": payload.get("numero_motor"),
            "numero_cuadro": payload.get("numero_cuadro"),
            "nro_dnrpa": payload.get("nro_dnrpa"),
            "nro_certificado": payload.get("nro_certificado"),
        }
        where = []
        params = {}
        for field, value in checks.items():
            if value:
                where.append(f"v.{field} = :{field}")
                params[field] = value
        if not where:
            return []

        rows = db.execute(
            text(
                f"""
                SELECT
                    v.id, v.marca, v.modelo, v.anio,
                    v.nro_certificado, v.nro_dnrpa, v.lca,
                    v.numero_cuadro, v.numero_motor,
                    v.precio_lista, v.color_id, c.nombre AS color_nombre,
                    v.estado_stock_id, v.estado_moto_id, v.observaciones
                FROM vehiculos v
                LEFT JOIN colores c ON c.id = v.color_id
                WHERE {" OR ".join(where)}
                ORDER BY v.id ASC
                """
            ),
            params,
        ).mappings().all()
        return [dict(r) for r in rows]

    def _diffs(self, existing: Dict[str, Any], payload: Dict[str, Any]) -> List[Dict[str, Any]]:
        diffs: List[Dict[str, Any]] = []
        for field in SYSTEM_FIELDS:
            new_value = payload.get(field)
            if _is_empty(new_value):
                continue
            old_value = existing.get(field)
            if field == "precio_lista":
                equal = _money_equal(old_value, new_value)
            else:
                equal = _norm_compare(old_value) == _norm_compare(new_value)
            if not equal:
                diffs.append(
                    {
                        "campo": field,
                        "label": FIELD_LABELS.get(field, field),
                        "sistema": old_value,
                        "excel": new_value,
                    }
                )
        return diffs

    def _build_update_payload(self, row: Dict[str, Any], *, only_empty: bool) -> Dict[str, Any]:
        existing = row.get("existing") or {}
        payload = row.get("payload") or {}
        update_data: Dict[str, Any] = {}
        for diff in row.get("diferencias") or []:
            field = diff.get("campo")
            if field not in SYSTEM_FIELDS:
                continue
            if only_empty and not _is_empty(existing.get(field)):
                continue
            value = payload.get(field)
            if not _is_empty(value):
                update_data[field] = value
        return update_data

    def _validate_create_payload(self, payload: Dict[str, Any], row: Dict[str, Any]) -> None:
        missing = self._create_payload_errors(payload, row)
        if missing:
            raise ValueError("Faltan datos requeridos para crear: " + ", ".join(missing))

    def _create_payload_errors(self, payload: Dict[str, Any], row: Dict[str, Any]) -> List[str]:
        missing = []
        for field in ("marca", "modelo", "nro_certificado", "numero_motor", "precio_lista", "color_id"):
            if field == "color_id" and _clean_str((row.get("excel") or {}).get("color_nombre")):
                continue
            if _is_empty(payload.get(field)):
                missing.append(FIELD_LABELS.get(field, field))
        return missing

    def _audit_action_for_import(self, action: str) -> str:
        return {
            "COMPLETAR_VACIOS": "IMPORT_CERT_COMPLETAR_VACIOS",
            "ACTUALIZAR_DIFERENCIAS": "IMPORT_CERT_ACTUALIZAR_DIF",
        }.get(action, "IMPORT_CERT_ACTUALIZAR")

    def _row_errors(self, payload: Dict[str, Any], excel: Dict[str, Any]) -> List[str]:
        errors = []
        if _is_empty(payload.get("marca")):
            errors.append("No se pudo detectar marca.")
        if _is_empty(payload.get("modelo")):
            errors.append("Falta modelo.")
        if _is_empty(payload.get("numero_motor")):
            errors.append("Falta numero de motor.")
        return errors

    def _build_observaciones(self, raw: Dict[str, Any]) -> str:
        parts = []
        ubicacion = _clean_str(raw.get("ubicacion"))
        if ubicacion:
            parts.append(ubicacion)
        return " | ".join(parts)

    def _crear_colores_faltantes_en_sesion(self, db: Session, rows: List[Dict[str, Any]]) -> None:
        selected_missing = self.colores_faltantes(
            [row for row in rows if row.get("accion") and row.get("accion") != "OMITIR"]
        )
        if not selected_missing:
            return

        color_ids: Dict[str, int] = {}
        for color in selected_missing:
            existing = db.execute(
                text(
                    """
                    SELECT id, nombre
                    FROM colores
                    WHERE UPPER(nombre) = :nombre
                    LIMIT 1
                    """
                ),
                {"nombre": color.upper()},
            ).mappings().first()
            if existing:
                color_ids[_norm_color(color)] = int(existing["id"])
                continue

            res = db.execute(
                text("INSERT INTO colores (nombre) VALUES (:nombre)"),
                {"nombre": color.upper()},
            )
            color_id = int(res.lastrowid or 0)
            color_ids[_norm_color(color)] = color_id
            self._audit.registrar(
                db,
                entidad="colores",
                entidad_id=color_id or None,
                accion="IMPORTACION_CERTIFICADOS_CREAR_COLOR",
                datos_nuevos={"nombre": color.upper()},
                contexto={"importacion": "certificados_avanzada"},
            )

        for row in rows:
            if not row.get("accion") or row.get("accion") == "OMITIR":
                continue
            excel = row.get("excel") or {}
            payload = row.get("payload") or {}
            color_name = _clean_str(excel.get("color_nombre"))
            if color_name and _is_empty(payload.get("color_id")):
                color_id = color_ids.get(_norm_color(color_name))
                if color_id:
                    payload["color_id"] = color_id

    def _color_map(self) -> Dict[str, Dict[str, Any]]:
        result = {}
        for color in self._catalogos.get_colores():
            name = color.get("nombre")
            result[_norm_color(name)] = color
            if _norm_color(name).endswith("o"):
                result[_norm_color(name)[:-1] + "a"] = color
            if _norm_color(name).endswith("a"):
                result[_norm_color(name)[:-1] + "o"] = color
        return result

    def _default_condicion_id(self) -> Optional[int]:
        condiciones = self._catalogos.get_condiciones()
        for item in condiciones:
            if "nueva" in _norm(item.get("nombre")):
                return item.get("id")
        return condiciones[0].get("id") if condiciones else None


def _clean_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm(value: Any) -> str:
    text_value = _clean_str(value).lower()
    text_value = unicodedata.normalize("NFKD", text_value)
    text_value = "".join(ch for ch in text_value if not unicodedata.combining(ch))
    text_value = text_value.replace("º", "").replace("°", "")
    text_value = re.sub(r"[^a-z0-9]+", " ", text_value)
    return re.sub(r"\s+", " ", text_value).strip()


def _norm_color(value: Any) -> str:
    return _norm(value)


def _norm_compare(value: Any) -> str:
    return _norm(value)


def _is_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _parse_int(value: Any) -> Optional[int]:
    text_value = _clean_str(value)
    if not text_value:
        return None
    match = re.search(r"\d{4}", text_value)
    if match:
        return int(match.group(0))
    return int(float(text_value)) if text_value.replace(".", "", 1).isdigit() else None


def _parse_money(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text_value = _clean_str(value)
    text_value = text_value.replace("$", "").replace(" ", "")
    if "," in text_value and "." in text_value:
        if text_value.rfind(".") > text_value.rfind(","):
            text_value = text_value.replace(",", "")
        else:
            text_value = text_value.replace(".", "").replace(",", ".")
    elif "," in text_value:
        text_value = text_value.replace(",", ".")
    try:
        return float(text_value)
    except Exception:
        return None


def _money_equal(a: Any, b: Any) -> bool:
    try:
        return round(float(a or 0), 2) == round(float(b or 0), 2)
    except Exception:
        return _norm_compare(a) == _norm_compare(b)


def _brand_from_text(value: Any, known_brands: Optional[set[str]] = None) -> str:
    norm = _norm(value).upper()
    for brand in sorted(known_brands or KNOWN_BRANDS, key=len, reverse=True):
        if brand in norm:
            return brand
    return ""
