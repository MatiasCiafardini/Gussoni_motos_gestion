from __future__ import annotations
from typing import Dict, List, Any

from io import BytesIO

from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill

from app.data.database import SessionLocal
from app.repositories.clientes_repository import ClientesRepository
from app.repositories.vehiculos_repository import VehiculosRepository
from app.services.catalogos_service import CatalogosService


# =====================================================
# Excepción de importación
# =====================================================

class ImportacionError(Exception):
    def __init__(self, errores: List[str]):
        super().__init__("Error en importación")
        self.errores = errores


# =====================================================
# Service principal
# =====================================================

class ImportacionDatosService:
    """
    Servicio batch de importación de datos usando XLSX.
    """

    TABLAS_DISPONIBLES = {
        "clientes": "Clientes",
        "vehiculos": "Vehículos",
    }

    # ==================================================
    # API pública
    # ==================================================

    def listar_tablas(self) -> List[Dict[str, str]]:
        return [{"key": k, "label": v} for k, v in self.TABLAS_DISPONIBLES.items()]

    def generar_plantilla(self, tabla: str) -> bytes:
        if tabla == "clientes":
            return self._plantilla_clientes()
        if tabla == "vehiculos":
            return self._plantilla_vehiculos()
        raise ValueError("Tabla no soportada")

    def importar_xlsx(self, tabla: str, file_bytes: bytes) -> Dict[str, Any]:
        rows = self._parse_xlsx(file_bytes)
        if tabla == "clientes":
            return self._importar_clientes(rows)
        if tabla == "vehiculos":
            return self._importar_vehiculos(rows)
        raise ValueError("Tabla no soportada")

    # ==================================================
    # Parse XLSX
    # ==================================================

    def _parse_xlsx(self, file_bytes: bytes) -> List[Dict[str, Any]]:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active

        headers = [c.value for c in ws[1]]
        if not headers:
            raise ValueError("El archivo no tiene encabezados.")

        rows: List[Dict[str, Any]] = []

        for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            data = {"__rownum__": excel_row}
            for i, value in enumerate(row):
                if i < len(headers):
                    data[headers[i]] = value
            rows.append(data)

        if not rows:
            raise ValueError("El archivo no contiene datos.")

        return rows

    # ==================================================
    # ---------------- CLIENTES ------------------------
    # ==================================================

    def _plantilla_clientes(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        fill = PatternFill("solid", fgColor="E5E7EB")
        bold = Font(bold=True)

        headers = [
            ("tipo_doc", True),
            ("nro_doc", True),
            ("nombre", True),
            ("apellido", True),
            ("telefono", False),
            ("email", False),
            ("direccion", False),
            ("estado", False),
            ("observaciones", False),
        ]

        for col, (name, req) in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=name)
            if req:
                c.font = bold
            c.fill = fill
            ws.column_dimensions[c.column_letter].width = 22

        ws.append([
            "DNI", "30111222", "Juan", "Pérez",
            "11-4444-5555", "juan@mail.com",
            "Av. Siempre Viva 123", "Activo",
            "Cliente de ejemplo",
        ])

        catalogos = CatalogosService()
        tipos = [t["codigo"] for t in catalogos.get_tipos_documento()]

        dv_tipo = DataValidation(
            type="list",
            formula1=f'"{",".join(tipos)}"',
            allow_blank=False
        )

        ws.add_data_validation(dv_tipo)
        dv_tipo.add("A2:A1000")

        dv_estado = DataValidation(type="list", formula1='"Activo,Inactivo"', allow_blank=True)
        ws.add_data_validation(dv_estado)
        dv_estado.add("H2:H1000")

        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    def _importar_clientes(self, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
        db: Session = SessionLocal()
        repo = ClientesRepository(db)
        catalogos = CatalogosService()

        tipos = catalogos.get_tipos_documento()
        codigo_map = {t["codigo"].upper(): t["id"] for t in tipos}
        id_map = {t["id"]: t for t in tipos}

        errores, insertados = [], 0

        try:
            for row in rows:
                fila = row["__rownum__"]
                try:
                    payload = self._validar_cliente(row, repo, codigo_map, id_map)
                    repo.create_cliente(payload)
                    insertados += 1
                except Exception as e:
                    errores.append(f"Fila {fila}: {e}")

            if errores:
                raise ImportacionError(errores)

            db.commit()
            return {"success": True, "insertados": insertados}

        except ImportacionError as e:
            db.rollback()
            return {"success": False, "errores": e.errores}
        finally:
            db.close()


    def _validar_cliente(
        self,
        row: Dict[str, Any],
        repo: ClientesRepository,
        codigo_map: Dict[str, int],
        id_map: Dict[int, Dict[str, Any]],
    ) -> Dict[str, Any]:

        def req(k):
            v = row.get(k)
            if v is None or str(v).strip() == "":
                raise ValueError(f"{k} es obligatorio")
            return str(v).strip()

        codigo = req("tipo_doc").upper()

        if codigo not in codigo_map:
            raise ValueError(f"Tipo de documento inválido: {codigo}")

        tipo_doc_id = codigo_map[codigo]
        nro = "".join(ch for ch in req("nro_doc") if ch.isdigit())

        tipo_obj = id_map[tipo_doc_id]
        codigo_real = tipo_obj["codigo"].upper()

        if codigo_real == "DNI" and len(nro) not in (7, 8):
            raise ValueError("DNI inválido")

        if codigo_real in ("CUIT", "CUIL") and len(nro) != 11:
            raise ValueError("CUIT/CUIL inválido")

        if repo.exists_by_doc(tipo_doc_id, nro):
            raise ValueError("Documento duplicado")

        payload = {
            "tipo_doc_id": tipo_doc_id,
            "nro_doc": nro,
            "nombre": req("nombre"),
            "apellido": req("apellido"),
            "estado_id": 10,
        }

        if row.get("estado"):
            payload["estado_id"] = 11 if str(row["estado"]).lower() == "inactivo" else 10

        for opt in ("telefono", "email", "direccion", "observaciones"):
            if row.get(opt):
                payload[opt] = str(row[opt]).strip()

        return payload


    # ==================================================
    # ---------------- VEHÍCULOS -----------------------
    # ==================================================

    def _plantilla_vehiculos(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Vehiculos"

        catalogos = CatalogosService()
        colores = [c["nombre"] for c in catalogos.get_colores()]
        estados = [e["nombre"] for e in catalogos.get_estados_stock()]
        condiciones = [c["nombre"] for c in catalogos.get_condiciones()]
        proveedores = [p["nombre"] for p in catalogos.get_proveedores()]

        headers = [
            ("marca", True),
            ("modelo", True),
            ("anio", False),
            ("nro_certificado", True),
            ("nro_dnrpa", True),
            ("numero_cuadro", False),
            ("numero_motor", True),
            ("precio_lista", False),
            ("color", True),
            ("estado_stock", True),
            ("condicion", True),
            ("proveedor", False),
            ("observaciones", False),
        ]

        bold = Font(bold=True)
        fill = PatternFill("solid", fgColor="E5E7EB")

        for col, (name, req) in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=name)
            if req:
                c.font = bold
            c.fill = fill
            ws.column_dimensions[c.column_letter].width = 22

        ws.append([
            "Honda", "CG Titan 150", 2024,
            "EJEMPLO-CERT", "EJEMPLO-DNRPA",
            "EJEMPLO-CUADRO", "EJEMPLO-MOTOR",
            0, "Negro", "Disponible", "Nueva", "", "Unidad 0 km",
        ])

        def dv_list(col, values, allow_blank):
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(values)}"',
                allow_blank=allow_blank
            )
            ws.add_data_validation(dv)
            dv.add(f"{col}2:{col}1000")

        dv_list("I", colores, False)
        dv_list("J", estados, False)
        dv_list("K", condiciones, False)
        dv_list("L", proveedores, True)

        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    def _importar_vehiculos(self, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
        db: Session = SessionLocal()
        repo = VehiculosRepository(db)
        cat = CatalogosService()

        color_map = {c["nombre"].lower(): c["id"] for c in cat.get_colores()}
        estado_map = {e["nombre"].lower(): e["id"] for e in cat.get_estados_stock()}
        cond_map = {c["nombre"].lower(): c["id"] for c in cat.get_condiciones()}
        prov_map = {p["nombre"].lower(): p["id"] for p in cat.get_proveedores()}

        errores, insertados = [], 0

        try:
            for row in rows:
                fila = row["__rownum__"]
                try:
                    def req(k):
                        v = row.get(k)
                        if v is None or str(v).strip() == "":
                            raise ValueError(f"{k} es obligatorio")
                        return str(v).strip()

                    # -------- Validaciones --------
                    anio = row.get("anio")
                    if anio not in (None, ""):
                        anio = int(anio)
                        if anio < 1900 or anio > 2100:
                            raise ValueError("Año inválido (1900–2100)")

                    precio = float(row.get("precio_lista") or 0)
                    if precio < 0:
                        raise ValueError("Precio inválido")

                    nro_motor = req("numero_motor")
                    nro_dnrpa = req("nro_dnrpa")
                    nro_cuadro = row.get("numero_cuadro")

                    if repo.exists_by_numero_motor(nro_motor):
                        raise ValueError("Número de motor duplicado")
                    if repo.exists_by_nro_dnrpa(nro_dnrpa):
                        raise ValueError("Número DNRPA duplicado")
                    if nro_cuadro and repo.exists_by_numero_cuadro(str(nro_cuadro)):
                        raise ValueError("Número de cuadro duplicado")

                    payload = {
                        "marca": req("marca"),
                        "modelo": req("modelo"),
                        "anio": anio,
                        "nro_certificado": req("nro_certificado"),
                        "nro_dnrpa": nro_dnrpa,
                        "numero_cuadro": nro_cuadro,
                        "numero_motor": nro_motor,
                        "precio_lista": precio,
                        "color_id": color_map[req("color").lower()],
                        "estado_stock_id": estado_map[req("estado_stock").lower()],
                        "estado_moto_id": cond_map[req("condicion").lower()],
                        "observaciones": row.get("observaciones"),
                    }

                    prov = row.get("proveedor")
                    if prov:
                        if str(prov).lower() not in prov_map:
                            raise ValueError("Proveedor inválido")
                        payload["proveedor_id"] = prov_map[str(prov).lower()]

                    repo.create_vehiculo(payload)
                    insertados += 1

                except Exception as e:
                    errores.append(f"Fila {fila}: {e}")

            if errores:
                raise ImportacionError(errores)

            db.commit()
            return {"success": True, "insertados": insertados}

        except ImportacionError as e:
            db.rollback()
            return {"success": False, "errores": e.errores}
        finally:
            db.close()
